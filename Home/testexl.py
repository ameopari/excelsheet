import mysql.connector 
from mysql.connector import Error
import pandas as pd


import os
import openpyxl

from pathlib import Path
try:
    connection = mysql.connector.connect(host="localhost", database = 'books',user="test", password="123456789")
    print('conn=========',connection)
    query = "Select * from events"
    print('111111111')
    cursor = connection.cursor()
    cursor.execute(query)
    records = cursor.fetchall()
    print("Total number of rows in table: ", cursor.rowcount)

    print("\nPrinting each row")
    for events in records:
        print("title = ", events[1],)
        print("start_date = ", events[2],)
        print("end_date = ", events[4],)

except mysql.connector.Error as e:
    print("Error reading data from MySQL table", e)
finally:
    if connection.is_connected():
        cursor.close()
        connection.close()
    print("MySQL connection is closed")
    # if  mysql.connector.connection():
    #     mysql.connector.close()
    #     cursor.close()
    #     print("MySQL connection is closed")
        






# # print(BASE_DIR.parent)

# # path = pd.read_excel (os.path.join(BASE_DIR.parent,'1.xls'),sheet_name="SalesOrders")
# path = pd.read_excel('/home/as/Desktop/Restframework/Restframework/Home/Pub_Program_2021.xlsx',sheet_name="Sheet1")
# print('path========',path)
# wb_obj = openpyxl.load_workbook('/home/as/Desktop/Restframework/Restframework/Home/Pub_Program_2021.xlsx')
# sheet_obj = wb_obj.active
# print('workkkkkkkkkkkkkkkkkkkkk')
# cell_obj = sheet_obj.cell(row = 1, column = 1)
 
# # Print value of cell object
# # using the value attribute
# print('yessssssssssssss',cell_obj.value)



# from openpyxl import load_workbook
# import pandas as pd

# wb_obj = load_workbook('/home/as/Downloads/Pub_Program_20211.xlsx')
# ws = wb_obj['Sheet1']
# sheet_obj = wb_obj.active

# data = ws.values
# # Get the first line in file as a header line
# # columns = next(data)[0:]
# # Create a DataFrame based on the second and subsequent lines of data
# # df = pd.DataFrame(data, columns=columns)
# def validateExcel(filename):

#     xls = xlrd.open_workbook(filename)  
#     # print('df====',df)

#     # print('df====',df)
#     for sheet in xls.sheets():
#         header=""

#         number_of_rows = sheet.nrows
#         number_of_columns = sheet.ncols
#         sheetname = sheet.name          

#         mylist = []

#         for row in range (1, number_of_rows):  
#             sublist = [sheet.cell_value(row, col) for col in range(0, number_of_columns)]

#             if sublist not in mylist:
#                 mylist.append(sublist)
#             print (mylist)

#             return mylist
#     # print('df====',df)
# print(sheet_obj.max_column) 
# max_col = sheet_obj.max_column
# for i in range(1,max_col+1):
#     cell_obj = sheet_obj.cell(row=1,column=i)


